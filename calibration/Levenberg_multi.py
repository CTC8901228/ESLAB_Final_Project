import numpy as np
from numpy import matrix as mat
from matplotlib import pyplot as plt
import random
import openpyxl
workbook = openpyxl.load_workbook('data/static_z.xlsx')
sheet = workbook.worksheets[0]
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始，跳过标题行
    data.append(row)
workbook.close

nz = sheet.max_row - 1 #總共資料數

workbookx = openpyxl.load_workbook('data/static_x.xlsx')
sheetx = workbookx.worksheets[0]
datax = []
for row in sheetx.iter_rows(min_row=2, values_only=True):  # 从第二行开始，跳过标题行
    datax.append(row)
workbookx.close

nx = sheetx.max_row - 1 #總共資料數

workbooky = openpyxl.load_workbook('data/static_y.xlsx')
sheety = workbooky.worksheets[0]
datay = []
for row in sheety.iter_rows(min_row=2, values_only=True):  # 从第二行开始，跳过标题行
    datay.append(row)
workbooky.close

ny = sheety.max_row - 1 #總共資料數

# #a1,b1,c1 = 1,3,2 #虛擬合之真實參數
# alpha_yz, alpha_zy, alpha_zx = 0, 0, 0
# #T_a = np.array([[1,-alpha_yz,alpha_zy],[0,1,-alpha_zx],[0,0,1]])
# s_ax, s_ay, s_az = 0, 0, 0 #acc scale bias
# #K_a = np.array([[s_ax,0,0],[0,s_ay,0],[0,0,s_az]])
# b_ax, b_ay, b_az = 0, 0, 0 #acc bias
#h = np.linspace(0,1,n) #產生噪音 在0~1之間產生n個點

# y = [np.exp(a1*i**2+b1*i+c1) + random.gauss(0,4) for i in h]
# y = mat(y)

###### calibration of acc by using Levenberg ########
print ("start calibrate acc")

def Func (theta_acc, input_a, axis): #input_a: 3*1 array
    T_a = np.array([[1,-theta_acc[0,0],theta_acc[1,0]],[0,1,-theta_acc[2,0]],[0,0,1]])
    K_a = np.array([[theta_acc[3,0],0,0],[0,theta_acc[4,0],0],[0,0,theta_acc[5,0]]])
    b_a = np.array([[theta_acc[6,0]],[theta_acc[7,0]],[theta_acc[8,0]]])
    h = (T_a.dot(K_a)).dot((input_a + b_a))
    # h = (h[0]**2+h[1]**2+h[2]**2)**0.5
    if axis == 2:
        Loss = abs(abs(h[0])-0)**2 + abs(abs(h[1])-0)**2 + abs(h[2]+9.81)**2
    if axis == 0:
        Loss = abs(h[0]+9.81)**2 + abs(abs(h[1])-0)**2 + abs(h[2]-0)**2
    if axis == 1:
        Loss = abs(abs(h[0])-0)**2 + abs(h[1]-9.81)**2 + abs(h[2])**2
    
    return Loss 

def Deriv (theta_acc,input_a,n,axis):
    x1 = theta_acc.copy()
    x2 = theta_acc.copy()
    x1[n,0] -= 0.00000001
    x2[n,0] += 0.00000001
    p1 = Func(x1,input_a,axis)
    p2 = Func(x2,input_a,axis)
    d = (p2-p1)*1.0/(0.00000002)
    return d

J = mat(np.zeros(((nx+ny+nz),9)))
fx = mat(np.zeros(((nx+ny+nz),1)))
fx_tmp = mat(np.zeros(((nx+ny+nz),1)))
#xk = mat([[0.8],[2.7],[1.5]])
theta_acc = np.array([[0],[0],[0],[1],[1],[1],[-0.3],[0.2],[0.01]]) # alpha , s , b
lase_mse = 0
step = 0
u, v = 1,2
conve = 1000

while (conve):

    mse , mse_tmp = 0,0
    step += 1
    for i in range((nx+ny+nz)):
        if i < nz :
            axis = 2
            input_a = np.array([[data[i][4]],[data[i][5]],[data[i][6]]]) #取data的加速度項static的 ax ay az
            # fx[i] = abs(Func(theta_acc,input_a)) - 9.81
            fx[i] = Func(theta_acc,input_a,axis) 
            # mse += fx[i,0]**2
            mse += fx[i,0]
            
            for j in range(9):
                J[i,j] = Deriv(theta_acc,input_a,j,axis)
        
        if i >= nz  and i < (nz+nx):
            axis = 0
            input_a = np.array([[datax[i-nz][4]],[datax[i-nz][5]],[datax[i-nz][6]]]) #取data的加速度項static的 ax ay az
            fx[i] = Func(theta_acc,input_a,axis) 
            # mse += fx[i,0]**2
            mse += fx[i,0]
            
            for j in range(9):
                J[i,j] = Deriv(theta_acc,input_a,j,axis)

        if i >= (nz+nx):
            axis = 1
            input_a = np.array([[datay[i-(nz+nx)][4]],[datay[i-(nz+nx)][5]],[datay[i-(nz+nx)][6]]]) #取data的加速度項static的 ax ay az
            fx[i] = Func(theta_acc,input_a,axis) 
            # mse += fx[i,0]**2
            mse += fx[i,0]
            
            for j in range(9):
                J[i,j] = Deriv(theta_acc,input_a,j,axis)
    
    mse /= (nx+ny+nz) #約束範圍

    H = J.T*J + u*np.eye(9)
    dx = -H.I * J.T*fx
    xk_tmp = theta_acc.copy()
    xk_tmp = xk_tmp + dx

    for i in range ((nx+ny+nz)):
        if i < nz :
            axis = 2
            input_a = np.array([[data[i][4]],[data[i][5]],[data[i][6]]])
            fx_tmp[i] = Func(xk_tmp,input_a,axis)
        if i >= nz  and i < (nz+nx) :
            axis = 0
            input_a = np.array([[datax[i-nz][4]],[datax[i-nz][5]],[datax[i-nz][6]]])
            fx_tmp[i] = Func(xk_tmp,input_a,axis)
        if i >= (nz+nx) :
            axis = 1
            input_a = np.array([[datay[i-(nz+nx)][4]],[datay[i-(nz+nx)][5]],[datay[i-(nz+nx)][6]]])
            fx_tmp[i] = Func(xk_tmp,input_a,axis)
        # mse_tmp += fx_tmp[i,0]**2
        mse_tmp += fx_tmp[i,0]
    mse_tmp /= (nx+ny+nz)

    q = (mse - mse_tmp)/((0.5*dx.T*(u*dx - J.T*fx))[0,0])

    if q > 0:
        s = 1.0/3.0
        v = 2
        mse = mse_tmp
        theta_acc = xk_tmp
        temp = 1 - pow(2*q-1,3)
        if s > temp:
            u = u*s
        else:
            u = u*temp
    else:
        u = u*v
        v = 2*v
        theta_acc = xk_tmp
    
    print ("step = %d, abs(mse-lase_mse) = %.8f, %.8f, %.8f)",step , abs(mse-lase_mse),abs(lase_mse),abs(mse))
    if abs(mse-lase_mse) < 0.00000001:
        break
    lase_mse = mse
    conve -= 1

print ("final parameter to calibrate acc: ")
print (theta_acc)
# print ("1: ",theta_acc[0,0])
# print ("2: ",theta_acc[1,0])


###### calibration of gyro by using Levenberg ########
print ("start calibrate gyro")
gx_sum,gy_sum,gz_sum = 0,0,0
gx_avg,gy_avg,gz_avg = 0,0,0
for i in range(nz):
    gx_sum += data[i][1]
    gy_sum += data[i][2]
    gz_sum += data[i][3]
gx_avg = gx_sum/nz # get the bias of the gyro
gy_avg = gy_sum/nz
gz_avg = gz_sum/nz

def Func_gyro (theta_gyro, input_g): #input_a: 3*1 array
    T_g = np.array([[1,-theta_gyro[0,0],theta_gyro[1,0]],[theta_gyro[2,0],1,-theta_gyro[3,0]],[-theta_gyro[4,0],theta_gyro[5,0],1]])
    K_g = np.array([[theta_gyro[6,0],0,0],[0,theta_gyro[7,0],0],[0,0,theta_gyro[8,0]]])
    b_g = np.array([[-gx_avg],[-gy_avg],[-gz_avg]])
    h = (T_g.dot(K_g)).dot((input_g +b_g ))
    # h = (h[0]**2+h[1]**2+h[2]**2)**0.5
    Loss = abs(h[0])**2 + abs(h[1])**2 + abs(h[2])**2
    return Loss

def Deriv_gyro (theta_gyro,input_g,n):
    x1 = theta_gyro.copy()
    x2 = theta_gyro.copy()
    x1[n,0] -= 0.000001
    x2[n,0] += 0.000001
    p1 = Func_gyro(x1,input_g)
    p2 = Func_gyro(x2,input_g)
    d = (p2-p1)*1.0/(0.000002)
    return d

J = mat(np.zeros((nz,9)))
fx = mat(np.zeros((nz,1)))
fx_tmp = mat(np.zeros((nz,1)))
#xk = mat([[0.8],[2.7],[1.5]])
theta_gyro = np.array([[0.01],[0.01],[0.01],[0],[0],[0],[1],[1],[1]]) # r_yz r_zy r_xz r_zx r_xy r_yx s_gx s_gy s_gz
lase_mse = 0
step = 0
u, v = 1,2
conve = 1000

while (conve):

    mse , mse_tmp = 0,0
    step += 1
    for i in range(nz):
        input_g = np.array([[data[i][1]],[data[i][2]],[data[i][3]]]) #取data的加速度項static的 ax ay az
        fx[i] = Func_gyro(theta_gyro,input_g) 
        # mse += fx[i,0]**2
        mse += fx[i,0]
        
        for j in range(9):
            J[i,j] = Deriv_gyro(theta_gyro,input_g,j)
    mse /= nz #約束範圍

    H = J.T*J + u*np.eye(9)
    dx = -H.I * J.T*fx
    xk_tmp = theta_gyro.copy()
    xk_tmp = xk_tmp + dx

    for i in range (nz):
        input_g = np.array([[data[i][1]],[data[i][2]],[data[i][3]]])
        fx_tmp[i] = Func_gyro(xk_tmp,input_g)
        # mse_tmp += fx_tmp[i,0]**2
        mse_tmp += fx_tmp[i,0]
    mse_tmp /= nz

    q = (mse - mse_tmp)/((0.5*dx.T*(u*dx - J.T*fx))[0,0])

    if q > 0:
        s = 1.0/3.0
        v = 2
        mse = mse_tmp
        theta_gyro = xk_tmp
        temp = 1 - pow(2*q-1,3)
        if s > temp:
            u = u*s
        else:
            u = u*temp
    else:
        u = u*v
        v = 2*v
        theta_gyro = xk_tmp
    
    print ("step = %d, abs(mse-lase_mse) = %.8f)",step , abs(mse-lase_mse))
    if abs(mse-lase_mse) < 1:
        break
    lase_mse = mse
    conve -= 1

print ("final parameter to calibrate gyro: ")
print (theta_gyro)



###### save in the parameter into .xlsx ########
workbook = openpyxl.Workbook()
sheet0 = workbook.worksheets[0]
listtitle = ["alpha_yx","alpha_zy", "alpha_zx", "scale_ax", "scale_ay", "scale_az", "bias_ax", "bias_ay","bias_az","r_yz", "r_zy", "r_xz","r_zx", "r_xy", "r_yx", "s_gx", "s_gy", "s_gz","gx_avg_bias(use plus)","gy_avg_bias","gz_avg_bias"]
sheet0.append (listtitle)
listtitle = [theta_acc[0,0],theta_acc[1,0],theta_acc[2,0],theta_acc[3,0],theta_acc[4,0],theta_acc[5,0],theta_acc[6,0],theta_acc[7,0],theta_acc[8,0],theta_gyro[0,0],theta_gyro[1,0],theta_gyro[2,0],theta_gyro[3,0],theta_gyro[4,0],theta_gyro[5,0],theta_gyro[6,0],theta_gyro[7,0],theta_gyro[8,0],gx_avg,gy_avg,gz_avg]
sheet0.append (listtitle)
workbook.save('data/calbration_para.xlsx')
workbook.close

############### calibration the acc #################
workbook = openpyxl.Workbook()
sheet1 = workbook.worksheets[0]
def Func1 (theta_acc, input_a): #input_a: 3*1 array
    T_a = np.array([[1,-theta_acc[0,0],theta_acc[1,0]],[0,1,-theta_acc[2,0]],[0,0,1]])
    K_a = np.array([[theta_acc[3,0],0,0],[0,theta_acc[4,0],0],[0,0,theta_acc[5,0]]])
    b_a = np.array([[theta_acc[6,0]],[theta_acc[7,0]],[theta_acc[8,0]]])
    h = (T_a.dot(K_a)).dot((input_a + b_a))
    return h
listtitle = ["ax_cali","ay_cali", "az_cali"]
sheet1.append (listtitle)

for i in range(nz):
    input_a = np.array([[data[i][4]],[data[i][5]],[data[i][6]]])
    acc_calibra = Func1 (theta_acc, input_a)
    listtitle = [acc_calibra[0,0],acc_calibra[1,0], acc_calibra[2,0]]
    sheet1.append (listtitle)

workbook.save('data/acc_calbration.xlsx')
workbook.close

workbook = openpyxl.Workbook()
sheet0 = workbook.worksheets[0]
listtitle = ["ax_cali","ay_cali", "az_cali"]
sheet0.append (listtitle)
for i in range(nx):
    input_a = np.array([[datax[i][4]],[datax[i][5]],[datax[i][6]]])
    acc_calibra = Func1 (theta_acc, input_a)
    listtitle = [acc_calibra[0,0],acc_calibra[1,0], acc_calibra[2,0]]
    sheet0.append (listtitle)

workbook.save('data/acc_calbration_1.xlsx')
workbook.close

workbook = openpyxl.Workbook()
sheet0 = workbook.worksheets[0]
listtitle = ["ax_cali","ay_cali", "az_cali"]
sheet0.append (listtitle)
for i in range(ny):
    input_a = np.array([[datay[i][4]],[datay[i][5]],[datay[i][6]]])
    acc_calibra = Func1 (theta_acc, input_a)
    listtitle = [acc_calibra[0,0],acc_calibra[1,0], acc_calibra[2,0]]
    sheet0.append (listtitle)
workbook.save('data/acc_calbration_y.xlsx')
workbook.close
############### calibration the gyr #################
workbook = openpyxl.Workbook()
sheet2 = workbook.worksheets[0]
def Func2 (theta_gyro, input_g): #input_a: 3*1 array
    T_g = np.array([[1,-theta_gyro[0,0],theta_gyro[1,0]],[theta_gyro[2,0],1,-theta_gyro[3,0]],[-theta_gyro[4,0],theta_gyro[5,0],1]])
    K_g = np.array([[theta_gyro[6,0],0,0],[0,theta_gyro[7,0],0],[0,0,theta_gyro[8,0]]])
    b_g = np.array([[-gx_avg],[-gy_avg],[-gz_avg]])
    h = (T_g.dot(K_g)).dot((input_g + b_g))
    return h
listtitle = ["gx_cali","gy_cali", "gz_cali"]
sheet2.append (listtitle)

for i in range(nz):
    input_g = np.array([[data[i][1]],[data[i][2]],[data[i][3]]])
    gyro_calibra = Func2 (theta_gyro, input_g)
    listtitle = [gyro_calibra[0,0],gyro_calibra[1,0], gyro_calibra[2,0]]
    sheet2.append (listtitle)

workbook.save('data/gyr_calbration.xlsx')
workbook.close