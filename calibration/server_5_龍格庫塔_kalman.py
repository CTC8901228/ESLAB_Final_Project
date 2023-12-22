import socket
import json
import numpy as np
import matplotlib.pyplot as plot
import math
from math import cos, sin, radians
import openpyxl
from collections import deque
from pykalman import KalmanFilter


HOST = '192.168.1.103' # IP address
PORT = 15000 # Port to listen on (use ports > 1023)
sampleFreq = 10 #Hz
linVel = np.array([0,0,0])
linPos = np.array([0,0,0])
acc_c = np.array([0,0,0])
axis_x = np.array([])
axis_y = np.array([])
axis_z = np.array([])
sp_1 = 0
asensitivity = 2048/2 #LSB/g 再除2的原因與靜止實驗結果有關
gsensitivity = 16.4 #LSB/(drg/s)
var = 0.0
deg_x, deg_y, deg_z = 0,0,0
newest_10gx = deque(maxlen=10)
newest_10gy = deque(maxlen=10)
newest_10gz = deque(maxlen=10)
newest_10ax = deque(maxlen=10)
newest_10ay = deque(maxlen=10)
newest_10az = deque(maxlen=10)
var = 0
w_1 = np.array([[0],[0],[0]]) 
q = np.array([[1],[0],[0],[0]]) #四元數 q0 q1 q2 q3 = w x y z

check = False ## 確認是否可以開始積分

##畫圖用
max_data_points = 30
plot_x = deque(maxlen=max_data_points)
plot_y = deque(maxlen=max_data_points)
plot_z = deque(maxlen=max_data_points)


############### Function of calibration ###################
#https://carlyleliu.github.io/2020/imu%E6%A0%A1%E5%87%86/

#get the calibration parameter
workbook = openpyxl.load_workbook('data/calbration_para.xlsx')
sheet = workbook.worksheets[0]
# alpha_yx	alpha_zy	alpha_zx	scale_ax	scale_ay	scale_az	bias_ax	bias_ay	bias_az	r_yz	r_zy	r_xz	r_zx	r_xy	r_yx	s_gx	s_gy	s_gz	gx_avg_bias(use plus)	gy_avg_bias	gz_avg_bias
cali_para = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始，跳过标题行
    cali_para.append(row)
workbook.close

# func of acc calibrate
def cali_acc (acc):
    K_a = np.array([[cali_para[0][3],0,0],[0,cali_para[0][4],0],[0,0,cali_para[0][5]]])
    T_a = np.array([[1,-cali_para[0][0],cali_para[0][1]],[0,1,-cali_para[0][2]],[0,0,1]])
    b_a = np.array([[cali_para[0][6]],[cali_para[0][7]],[cali_para[0][8]]])
    h = (T_a.dot(K_a)).dot((acc + b_a))
    return h

# func of gyro calibrate
def cali_gyro (gyro):
    T_g = np.array([[1,-cali_para[0][9],cali_para[0][10]],[cali_para[0][11],1,-cali_para[0][12]],[-cali_para[0][13],cali_para[0][14],1]])
    K_g = np.array([[cali_para[0][15],0,0],[0,cali_para[0][16],0],[0,0,cali_para[0][17]]])
    b_g = np.array([[-cali_para[0][18]],[-cali_para[0][19]],[-cali_para[0][20]]])
    h = (T_g.dot(K_g)).dot((gyro +b_g ))
    return h

############### Function from "Oscillatory-Motion-Tracking-With-x-IMU" ###################
#https://github.com/xioTechnologies/Oscillatory-Motion-Tracking-With-x-IMU

def MahonyAHRSupdateIMU(gx, gy, gz, ax, ay, az,sampleFreq):
    global twoKp, twoKi, q0, q1, q2, q3, integralFBx, integralFBy, integralFBz

    twoKpDef = 2.0 * 0.5  # 2 * proportional gain
    twoKiDef = 2.0 * 0.0  # 2 * integral gain

    # Variable definitions
    twoKp = twoKpDef  # 2 * proportional gain (Kp)
    twoKi = twoKiDef
    q0, q1, q2, q3 = 1.0, 0.0, 0.0, 0.0
    integralFBx, integralFBy, integralFBz = 0.0, 0.0, 0.0
    
    recipNorm = 0.0
    halfvx, halfvy, halfvz = 0.0, 0.0, 0.0
    halfex, halfey, halfez = 0.0, 0.0, 0.0
    qa, qb, qc = 0.0, 0.0, 0.0

    # Compute feedback only if accelerometer measurement valid (avoids NaN in accelerometer normalization)
    if not (ax == 0.0 and ay == 0.0 and az == 0.0):

        # Normalise accelerometer measurement
        recipNorm = 1.0 / math.sqrt(ax * ax + ay * ay + az * az)
        ax *= recipNorm
        ay *= recipNorm
        az *= recipNorm

        # Estimated direction of gravity and vector perpendicular to magnetic flux
        halfvx = q1 * q3 - q0 * q2
        halfvy = q0 * q1 + q2 * q3
        halfvz = q0 * q0 - 0.5 + q3 * q3
        # halfvx = 2*(q1 * q3 - q0 * q2)
        # halfvy = 2*(q0 * q1 + q2 * q3)
        # halfvz = q0 * q0 - q1*q1-q2*q2 + q3 * q3

        # Error is sum of cross product between estimated and measured direction of gravity
        halfex = (ay * halfvz - az * halfvy)
        halfey = (az * halfvx - ax * halfvz)
        halfez = (ax * halfvy - ay * halfvx)

        # Compute and apply integral feedback if enabled
        if twoKi > 0.0:
            integralFBx += twoKi * halfex * (1.0 / sampleFreq)  # integral error scaled by Ki
            integralFBy += twoKi * halfey * (1.0 / sampleFreq)
            integralFBz += twoKi * halfez * (1.0 / sampleFreq)
            gx += integralFBx  # apply integral feedback
            gy += integralFBy
            gz += integralFBz
        else:
            integralFBx = 0.0  # prevent integral windup
            integralFBy = 0.0
            integralFBz = 0.0

        # Apply proportional feedback
        gx += twoKp * halfex
        gy += twoKp * halfey
        gz += twoKp * halfez

    # Integrate rate of change of quaternion
    gx *= (0.5 * (1.0 / sampleFreq))  # pre-multiply common factors
    gy *= (0.5 * (1.0 / sampleFreq))
    gz *= (0.5 * (1.0 / sampleFreq))
    qa, qb, qc = q0, q1, q2
    q0 += (-qb * gx - qc * gy - q3 * gz)
    q1 += (qa * gx + qc * gz - q3 * gy)
    q2 += (qa * gy - qb * gz + q3 * gx)
    q3 += (qa * gz + qb * gy - qc * gx)

    # Normalise quaternion
    recipNorm = 1.0 / math.sqrt(q0 * q0 + q1 * q1 + q2 * q2 + q3 * q3)
    q0 *= recipNorm
    q1 *= recipNorm
    q2 *= recipNorm
    q3 *= recipNorm
    
    return q0,q1,q2,q3

def quatern2rotMat(q):
    R = np.zeros((len(q), 3, 3))

    R1 = 2 * (q[0]**2) - 1 + 2 * (q[1]**2)
    R2 = 2 * (q[1] * q[2] + q[0] * q[3])
    R3 = 2 * (q[1] * q[3] - q[0] * q[2])

    R4 = 2 * (q[1] * q[2] - q[0] * q[3])
    R5 = 2 * (q[0]**2) - 1 + 2 * (q[2]**2)
    R6 = 2 * (q[2] * q[3] + q[0] * q[1])

    R7 = 2 * (q[1] * q[3] + q[0] * q[2])
    R8 = 2 * (q[2] * q[3] - q[0] * q[1])
    R9 = 2 * (q[0]**2) - 1 + 2 * (q[3]**2)
    R = np.array([[R1,R2,R3],[R4,R5,R6],[R7,R8,R9]])
    return R


#################### 龍格庫塔 ######################
def Quaternion_diff(w,q): #w : 角速度(wx wy wz), q: 四元數(q0 q1 q2 q3) 4*1
    wx = w[0][0]
    wy = w[1][0]
    wz = w[2][0]
    W = np.array ([[0, -wx, -wy, -wz],[wx, 0, -wz, -wy],[wy, -wz, 0, wx],[wz, wy, -wx, 0]])
    
    return 0.5*W.dot(q)

def runge_kutta (w,w_1,q,t):
    # y is the initial value for y
    # x is the initial value for x
    # dx is the time step in x
    # f is derivative of function y(t)
    s1 = t * Quaternion_diff(w_1,q)
    s2 = t * Quaternion_diff(0.5*(w+w_1), q+0.5*t*s1)
    s3 = t * Quaternion_diff(0.5*(w+w_1), q+0.5*t*s2)
    s4 = t * Quaternion_diff(w, q+t*s3)
    q_new =  q + (s1 + 2*s2 + 2*s3 + s4)/6
    q_norm = q_new/np.linalg.norm(q_new)    
    return q_norm # w x y z = q0 q1 q2 q3

def quart_to_rpy(q):
    w = q[0][0]
    x = q[1][0]
    y = q[2][0]
    z = q[3][0]
    roll = math.atan2(2 * (w * x + y * z), 1 - 2 * (x * x + y * y))
    pitch = math.asin(2 * (w * y - x * z))
    yaw = math.atan2(2 * (w * z + x * y), 1 - 2 * (z * z + y * y))
    return roll, pitch, yaw

################### remove bias g ######################
bias_g = np.array([[0.0],[0.0],[0.0]]) 
sum_g = np.array([[0.0],[0.0],[0.0]])
num_bias = 0 


################### kalman ######################
ax_var = 9.70591E-05
ay_var = 7.12466E-05
az_var = 0.000182277
ax_1_means = np.zeros((1, 3))
ax_1_covariances = np.zeros((1, 3, 3))
ay_1_means = np.zeros((1, 3))
ay_1_covariances = np.zeros((1, 3, 3))
az_1_means = np.zeros((1, 3))
az_1_covariances = np.zeros((1, 3, 3))

kal_ax = np.zeros((1, 3))
kal_ay = np.zeros((1, 3))
kal_az = np.zeros((1, 3))

def kal_a (a,a_1,cov,axis,t,dt):
    # transition_matrix  
    F = [[1, dt, 0.5*dt**2], 
        [0,  1,       dt],
        [0,  0,        1]]

    # observation_matrix   
    H = [0, 0, 1]

    # transition_covariance 
    #透過轉移協方差 Q 可以描述系統雜訊。矩陣值越小，系統雜訊越小。濾波器將變得更硬並且估計將被延遲。與新測量相比，系統過去的權重將會更高。否則，過濾器將更加靈活，並且會對每個新測量做出強烈反應。
    Q = [[0.5,    0,      0], 
        [  0,  0.1,      0],
        [  0,    0,  10e-4]]
    #initial_state_mean
    X0 = [0,0,a]

    if axis == 0:
        R = ax_var
        # initial_state_covariance
        P0 = [[  0,    0,               0], 
            [  0,    0,               0],
            [  0,    0,   ax_var]]
        
    if axis == 1:
        R = ay_var
        # initial_state_covariance
        P0 = [[  0,    0,               0], 
            [  0,    0,               0],
            [  0,    0,   ay_var]]
        
    if axis == 2:
        R = az_var
        # initial_state_covariance
        P0 = [[  0,    0,               0], 
            [  0,    0,               0],
            [  0,    0,   az_var]]

    kf = KalmanFilter(transition_matrices = F, 
                  observation_matrices = H, 
                  transition_covariance = Q, 
                  observation_covariance = R, 
                  initial_state_mean = X0, 
                  initial_state_covariance = P0)    
    if t == 0:

        a = X0
        cov = P0
        # if axis == 1:
        #     a = X0
        #     cov = P0
        # if axis == 2:
        #     a = X0
        #     cov = P0
    else:

        a, cov = (
            kf.filter_update(
                a_1,
                cov,
                a
            )
        )
    return a, cov

gx_var = 0.069996291
gy_var = 0.090028609
gz_var = 0.025021423
gx_1_means = np.zeros((1, 2))
gx_1_covariances = np.zeros((1, 2,2))
gy_1_means = np.zeros((1, 2))
gy_1_covariances = np.zeros((1, 2, 2))
gz_1_means = np.zeros((1,2))
gz_1_covariances = np.zeros((1, 2, 2))

kal_gx = np.zeros((1, 2))
kal_gy = np.zeros((1, 2))
kal_gz = np.zeros((1, 2))

def kal_g (g,g_1,cov,axis,t,dt):
    # transition_matrix  
    F = [[ 1,       dt],
        [0,        1]]

    # observation_matrix   
    H = [ 0, 1]

    
    #initial_state_mean
    X0 = [0,g]

    if axis == 0:
        R = gx_var
        # initial_state_covariance
        P0 = [[ 0,       0],
            [  0,   gx_var]]
        # transition_covariance 
        Q = [[ 0.01,      0],
            [   0,  10e-4]]
        
    if axis == 1:
        R = gy_var
        # initial_state_covariance
        P0 = [[ 0,     0],
            [ 0,   gy_var]]
        # transition_covariance 
        Q = [[ 0.05,      0],
            [   0,  10e-4]]
        
    if axis == 2:
        R = gz_var
        # initial_state_covariance
        P0 = [[0,   0],
            [0,   gz_var]]
        # transition_covariance 
        Q = [[ 0.05,      0],
            [   0,  10e-4]]

    kf = KalmanFilter(transition_matrices = F, 
                  observation_matrices = H, 
                  transition_covariance = Q, 
                  observation_covariance = R, 
                  initial_state_mean = X0, 
                  initial_state_covariance = P0)    
    if t == 0:

        g = X0
        cov = P0

    else:

        g, cov = (
            kf.filter_update(
                g_1,
                cov,
                g
            )
        )
    return g, cov



def plot_3axis (a,b,c):
    plot_x.append(a)
    plot_y.append(b)
    plot_z.append(c)
    plot.ion()
    plot.clf()
    plot.plot(plot_x)
    plot.plot(plot_y)
    plot.plot(plot_z)
    plot.pause(0.0000001)
    plot.show()



#################### start socket ######################
workbook1 = openpyxl.Workbook()
sheet1 = workbook1.worksheets[0]
# fig, ax = plot.subplots(2,2)

t=0
with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
    s.bind((HOST, PORT))
    s.listen()
    print("Starting server at: ", (HOST, PORT))
    conn, addr = s.accept()
    with conn:
        print("Connected at", addr)
        while True:
            data = conn.recv(1024*1024*8).decode('utf-8')
            #print("Received from socket server:", data)
            if (data.count('{') != 1):
                # Incomplete data are received.
                choose = 0
                buffer_data = data.split('}')
                while buffer_data[choose][0] != '{':
                    choose += 1
                data = buffer_data[choose] + '}'
            obj = json.loads(data)

            sp = obj['sp']
            gx = -(obj['gx'])/gsensitivity         #degree/s
            gy = -(obj['gy'])/gsensitivity
            gz = (obj['gz'])/gsensitivity
            ax = obj['ax']*9.81/asensitivity      #unit : m/s^2
            ay = obj['ay']*9.81/asensitivity
            az = -obj['az']*9.81/asensitivity
            sampleFreq = 1/(sp-sp_1)
            sp_1 = sp
            #print (gx, gy, gz)
            # print (ax, ay, az)
            # acc = np.array([ax,ay,az]) # 負號為調整成正確坐標系
            # gyr = np.array([gx,gy,gz])
            ##### calibration #####


            gyr = np.array([[gx],[gy],[gz]]) #degree/s
            acc = np.array([[ax],[ay],[az]]) #unit : m/s^2
            acc = cali_acc (acc) # 3*1
            print (acc)
            print (np.size(acc))

            acc = acc.T [0]
            gyr_d = cali_gyro (gyr) # 3*1

            ####### remove_bias(gyr_d)
            sum_g += gyr_d
            num_bias += 1
            if num_bias == 50:
                bias_g = sum_g / 50
                print (bias_g[0][0], bias_g[1][0], bias_g[2][0])
                print("can start moving")
                check = True
            gyr_d -= bias_g
            #print (num_bias, gyr_d[0])
            #plot_3axis(gyr_d[0][0],gyr_d[1][0],gyr_d[2][0])

            #plot_3axis(gyr[0][0]-cali_para[0][18],gyr[1][0]-cali_para[0][19],gyr[2][0]-cali_para[0][20])
            #### 單位轉換
            #gyr = gyr_d.T[0] #unit degree /s
            gyr_r = gyr_d * (math.pi/180) #unit radius /s
            

            # (q) = MahonyAHRSupdateIMU(gyr[0], gyr[1], gyr[2], (acc[0]/9.81), (acc[1]/9.81), (acc[2]/9.81),sampleFreq) # input unit shoud be radius and g
            
            # R = quatern2rotMat(q).T
            # tcAcc = np.dot(R,(acc.T/9.81)) # Calculate 'tilt-compensated' accelerometer
            # linAcc = tcAcc - np.array([0,0,-1])# Calculate linear acceleration in Earth frame (subtracting gravity)
            # linAcc *= 9.81 # unit m/s^2
            # linAcc[2] += 9.81
            #acc[2] -= 9.81
            #print (linAcc)

            ###### 龍格庫塔 用徑度
            # q = runge_kutta(gyr_r, w_1, q, 1/sampleFreq)
            # w_1 = gyr
            # roll, pitch, yaw = quart_to_rpy(q)
            # print ("%.4f  %.4f  %.4f" %(roll/(2*math.pi), pitch/(2*math.pi), yaw/(2*math.pi)))
            # plot_3axis (roll/(2*math.pi),pitch/(2*math.pi),yaw/(2*math.pi))

            ###### linear integrate the degree (和龍格庫塔 則一使用)
            newest_10gx.append(gyr_d[0][0])
            newest_10gy.append(gyr_d[1][0])
            newest_10gz.append(gyr_d[2][0])
            newest_10ax.append(acc[0])
            newest_10ay.append(acc[1])
            newest_10az.append(acc[2])

            dt = 1/sampleFreq

            kal_gx, gx_1_covariances = kal_g (gyr_d[0][0],kal_gx,gx_1_covariances,0,t,dt) # kal_g = [degree, degree/s]
            kal_gy, gy_1_covariances = kal_g (gyr_d[1][0],kal_gy,gy_1_covariances,1,t,dt)
            kal_gz, gz_1_covariances = kal_g (gyr_d[2][0],kal_gz,gz_1_covariances,2,t,dt)
            #print (kal_ax[2] ,kal_ay[2], kal_az[2])
            #plot_3axis(kal_ax[2] ,kal_ay[2], kal_az[2])
            #plot_3axis(kal_gx[0] ,kal_gy[0], kal_gz[0])
            print (sp ,kal_gx[0]/7.62831 ,kal_gy[0]/7.588, kal_gz[0]/7.72638)
            deg_x = kal_gx[0]/7.62831
            deg_y = kal_gy[0]/7.588
            deg_z = kal_gz[0]/7.72638

            if (np.var(newest_10ax) > 0.0005 or np.var(newest_10ay) > 0.0005 or np.var(newest_10az) > 0.0005):
                # if len (newest_10gx) == 10:
                #     if np.var(newest_10gx)>80:
                #         deg_x = deg_x + gyr_d[0][0]/sampleFreq
                #     if np.var(newest_10gy)>80:
                #         deg_y = deg_y + gyr_d[1][0]/sampleFreq
                #     if np.var(newest_10gz)>80 and np.var(newest_10gx)<1000:
                #         deg_z = deg_z + gyr_d[2][0]/sampleFreq
                # deg_x = deg_x + gyr_d[0][0]/sampleFreq
                # deg_y = deg_y + gyr_d[1][0]/sampleFreq 
                # deg_z = deg_z + gyr_d[2][0]/sampleFreq      
                C_b = np.array([[-sin(radians(deg_y/3.696))*9.81],
                                [cos(radians(deg_y/3.696))*sin(radians(deg_x/3.85836))*9.81],
                                [cos(radians(deg_y/3.696))*cos(radians(deg_x/3.85836))*9.81]])  ##重力轉換矩陣
                

                acc_c = acc + C_b.T
                #print (" %.6f %.6f %.6f" %(acc_c [0][0] , acc_c[1][0],acc_c[2][0] ))
                #print (acc , deg_x,  deg_y, deg_z)

                acc_c = acc_c[0]

                ######## kalman 
                
                ### 被角度校正過重力
                # kal_ax, ax_1_covariances = kal_a (acc_c[0],kal_ax,ax_1_covariances,0,t,dt)  # kal_a = [position, velocity, accelaration]
                # kal_ay, ay_1_covariances = kal_a (acc_c[1],kal_ay,ay_1_covariances,0,t,dt)
                # kal_az, az_1_covariances = kal_a (acc_c[2],kal_az,az_1_covariances,0,t,dt)

                ### 無校正過重力
                kal_ax, ax_1_covariances = kal_a (acc[0],kal_ax,ax_1_covariances,0,t,dt)
                kal_ay, ay_1_covariances = kal_a (acc[1],kal_ay,ay_1_covariances,0,t,dt)
                kal_az, az_1_covariances = kal_a (acc[2],kal_az,az_1_covariances,0,t,dt)


                # kal_gx, gx_1_covariances = kal_g (gyr_d[0][0],kal_gx,gx_1_covariances,0,t,dt) #[position, velocity, acceleration]
                # kal_gy, gy_1_covariances = kal_g (gyr_d[1][0],kal_gy,gy_1_covariances,0,t,dt)
                # kal_gz, gz_1_covariances = kal_g (gyr_d[2][0],kal_gz,gz_1_covariances,0,t,dt)
                # #print (kal_ax[2] ,kal_ay[2], kal_az[2])
                # #plot_3axis(kal_ax[2] ,kal_ay[2], kal_az[2])


            
                #print ("x drg: %.2f, %.4f, y drg: %.2f, %.4f, z drg: %.4f, %.4f" %(deg_x/(360*3.85836),np.var(newest_10gx), (deg_y/(360*3.696)),np.var(newest_10gy), deg_z/(360*3.7285),np.var(newest_10gz)))
              
                


                linVel = linVel + acc_c/sampleFreq

                linPos = linPos + linVel/sampleFreq
                linPos = linPos[0]

                ### 蒐集所有數據用
                if t == 0:
                    axis_x = kal_ax[0]
                    axis_y = kal_ay[0]
                    axis_z = kal_az[0]
                else:
                    axis_x = np.vstack((axis_x, kal_ax[0]))
                    axis_y = np.vstack((axis_y, kal_ay[0]))
                    axis_z = np.vstack((axis_z, kal_az[0]))
                    print("k")

                t = t+1
                #print (linPos[0],linPos[1],linPos[2], acc_c)
            #print(acc)


            
            #plot.scatter(sp, acc[0], c='blue') # x, y, z, gx, gy, gz
            #plot.xlabel("sample num, x")
            # # plot.scatter(t, gy, c='blue') # x, y, z, gx, gy, gz
            # # plot.xlabel("sample num, gy")
            # # plot.scatter(t, gz, c='blue') # x, y, z, gx, gy, gz
            # # plot.xlabel("sample num, gz")
            # t+=1
            #plot.pause(0.0001)

            #### 畫多圖 ####
            # plot.scatter(sp,linAcc[0],c='r',label="x")
            # plot.scatter(sp,linAcc[1],c='g',label="y")
            # plot.scatter(sp,linAcc[2],c='b',label="z")

            # plot.pause(0.0001)
            ##########

            #print (" %.6f %.6f %.6f" %(ax , ay,az ))
            # print (192.168.178.95,ay,az)
            # print (linPos)
            # listtitle = [sp ,linPos[0],linPos[1], linPos[2]]
            # sheet1.append (listtitle)
            
#             if sp >15:
#                 break
# # workbook1.save('data/move.xlsx')

# # 建立3D圖形
# fig = plot.figure()
# ax = fig.add_subplot(111, projection='3d')

# # 繪製3D點雲
# ax.scatter(axis_x, axis_y, axis_z)

# # 設置坐標軸標籤
# ax.set_xlabel('X Label')
# ax.set_ylabel('Y Label')
# ax.set_zlabel('Z Label')

# # 顯示圖形
# plot.show()